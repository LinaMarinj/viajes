import streamlit as st
import pandas as pd
import os
import plotly.express as px
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Configuración del estilo
st.markdown(
    """
    <style>
        .withScreencast > div {
            background-image: url('http://alavion.like-themes.com/wp-content/uploads/2018/02/SLIDE_02.jpg');
            background-size: cover;
            background-position: center;
            background-attachment: fixed;
        }
        
        .withScreencast header {
            background: none !important;
        }
        
        .stForm {
            background-color: rgba(255, 255, 255, 0.8);
            border-radius: 10px;
            padding: 20px;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Bienvenido(a) ¡A Viajar! ✈️")

# Inicializar variables de sesión
if "reservar" not in st.session_state:
    st.session_state.reservar = False
if "comprar" not in st.session_state:
    st.session_state.comprar = False


# Función para guardar en Excel
def guardar_reserva(datos):
    try:
        filename = "reservas_viajes_completas.xlsx"

        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            headers = [
                "Destino",
                "Personas",
                "Fecha Viaje",
                "Fecha Registro",
                "Cédula",
                "Nombre",
                "Apellido",
                "Teléfono",
                "Correo",
            ]
            ws.append(headers)
        else:
            wb = load_workbook(filename)
            ws = wb.active

        fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M")
        nueva_fila = [
            datos["destino"],
            datos["personas"],
            datos["fecha"].strftime("%Y-%m-%d"),
            fecha_registro,
            str(datos["cedula"]),
            datos["nombre"],
            datos["apellido"],
            str(datos["telefono"]),
            datos["correo"],
        ]

        ws.append(nueva_fila)
        wb.save(filename)
        return True
    except Exception as e:
        st.error(f"Error: {str(e)}")
        return False


# Formulario de reserva
with st.form("reserva_form"):
    col1, col2 = st.columns(2)

    with col1:
        destino = st.selectbox(
            "Destino",
            ["Selecciona un destino", "España", "Dubai", "Francia", "Italia", "Brasil"],
        )

    with col2:
        personas = st.number_input(
            "Personas a viajar", min_value=1, max_value=10, help="Número de viajeros"
        )

    fecha = st.date_input(
        "Fecha del viaje", min_value=datetime.today(), format="DD/MM/YYYY"
    )

    reservar = st.form_submit_button("Reservar")

if reservar:
    if destino == "Selecciona un destino":
        st.error("¡Por favor selecciona un destino válido!")
    else:
        st.session_state.reservar = True
        st.session_state.destino = destino
        st.session_state.personas = personas
        st.session_state.fecha = fecha

# Segundo formulario de datos personales
if st.session_state.reservar and not st.session_state.comprar:
    st.subheader("Información del comprador")
    with st.form("comprador_form"):
        cedula = st.text_input(
            "Cédula",
            placeholder="Ej: 1234567890",
            help="Solo números sin puntos ni guiones",
            key="cedula_input",
        )
        telefono = st.text_input(
            "Teléfono",
            placeholder="Ej: 3101234567",
            help="Solo números sin espacios ni guiones",
            key="telefono_input",
        )
        nombre = st.text_input(
            "Nombre",
            placeholder="Tu nombre",
            help="Ingresa tu nombre completo",
            key="nombre_input",
        )
        apellido = st.text_input(
            "Apellido",
            placeholder="Tu apellido",
            help="Ingresa tu apellido completo",
            key="apellido_input",
        )
        correo = st.text_input(
            "Correo Electrónico",
            placeholder="tucorreo@gmail.com",
            help="Ingresa un correo válido",
            key="correo_input",
        )

        comprar = st.form_submit_button("Confirmar compra")

    if comprar:
        errores = []
        if not cedula.isdigit():
            errores.append("La cédula debe contener solo números")
        if not telefono.isdigit():
            errores.append("El teléfono debe contener solo números")
        if not nombre.strip():
            errores.append("El nombre es obligatorio")
        if not apellido.strip():
            errores.append("El apellido es obligatorio")
        if "@" not in correo or "." not in correo:
            errores.append("Ingresa un correo electrónico válido")

        if errores:
            for error in errores:
                st.error(error)
        else:
            datos_reserva = {
                "destino": st.session_state.destino,
                "personas": st.session_state.personas,
                "fecha": st.session_state.fecha,
                "cedula": cedula,
                "nombre": nombre,
                "apellido": apellido,
                "telefono": telefono,
                "correo": correo,
            }

            if guardar_reserva(datos_reserva):
                st.session_state.comprar = True
                st.session_state.cedula = cedula
                st.session_state.nombre = nombre
                st.session_state.apellido = apellido
                st.session_state.telefono = telefono
                st.session_state.correo = correo
            else:
                st.error("Error al procesar la reserva")

# Mostrar confirmación final
if st.session_state.comprar:
    st.balloons()
    st.success("¡Reserva realizada con éxito! ✅")
    st.write("---")
    st.subheader("Detalles completos de la reserva:")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("*Información del viaje:*")
        st.write(f"- Destino: {st.session_state.destino}")
        st.write(f"- Viajeros: {st.session_state.personas} personas")
        st.write(f"- Fecha: {st.session_state.fecha.strftime('%d/%m/%Y')}")
        st.markdown("</div>", unsafe_allow_html=True)

    with col2:
        st.markdown("*Información personal:*")
        st.write(f"- Cédula: {st.session_state.cedula}")
        st.write(f"- Nombre: {st.session_state.nombre}")
        st.write(f"- Apellido: {st.session_state.apellido}")
        st.write(f"- Teléfono: {st.session_state.telefono}")
        st.write(f"- Correo: {st.session_state.correo}")
        st.markdown("</div>", unsafe_allow_html=True)

    filename = "reservas_viajes_completas.xlsx"

    if os.path.exists(filename):
        df = pd.read_excel(filename)

        # st.write("Datos cargados desde Excel:", df)

        if not df.empty:
            st.subheader("Historial de Reservas 📊")
            df["Fecha Viaje"] = pd.to_datetime(df["Fecha Viaje"])

            reservas_por_destino = (
                df.groupby(["Fecha Viaje", "Destino"])["Personas"].sum().reset_index()
            )

            fig = px.line(
                reservas_por_destino,
                x="Fecha Viaje",
                y="Personas",
                color="Destino",
                markers=True,
                title="",
            )

            st.plotly_chart(fig)

    st.write("")
    st.write("---")
    st.write("")

    # Botón de "Nueva Reserva"
    if st.button("🔄 Nueva Reserva", help="Haz clic para realizar una nueva reserva"):
        st.session_state.reservar = False
        st.session_state.comprar = False
        st.rerun()
